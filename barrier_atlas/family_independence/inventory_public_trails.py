#!/usr/bin/env python3
from __future__ import annotations
import argparse,hashlib,json
from pathlib import Path

TEXT_EXT={'.md','.txt','.csv','.tsv','.json','.yaml','.yml','.dat','.inp','.out'}

def sha256(p:Path):
    h=hashlib.sha256()
    with p.open('rb') as f:
        for b in iter(lambda:f.read(1<<20),b''): h.update(b)
    return h.hexdigest()

def write(path,obj):
    p=Path(path); p.parent.mkdir(parents=True,exist_ok=True)
    p.write_text(json.dumps(obj,indent=2,sort_keys=True)+'\n')

def manifest(root:Path):
    rows=[]
    for p in sorted(x for x in root.rglob('*') if x.is_file()):
        rows.append({'path':str(p.relative_to(root)),'size_bytes':p.stat().st_size,'sha256':sha256(p)})
    return rows

def text_excerpt(p:Path,limit=12000):
    try: return p.read_text(errors='replace')[:limit]
    except Exception: return ''

def trypsin(root:Path):
    files=manifest(root)
    key=[x for x in files if x['path'].startswith(('FES/','RATE/'))]
    excerpts={}
    for rel in ['README.md','FES/plumed.dat','FES/production.mdp','RATE/plumed.dat']:
        p=root/rel
        if p.is_file(): excerpts[rel]=text_excerpt(p)
    return {
        'schema':'family-independence-trypsin-inventory-v0.1',
        'role':'candidate_generation_only',
        'file_count':len(files),
        'fes_rate_file_count':len(key),
        'files':key,
        'text_excerpts':excerpts,
        'admission_decision':'NOT_EVALUATED'
    }

def enzyme_db(root:Path):
    files=manifest(root)
    systems=[]
    proteins=root/'proteins'
    if proteins.is_dir():
        for d in sorted(x for x in proteins.iterdir() if x.is_dir()):
            readmes=list(d.glob('readme*'))+list(d.glob('README*'))
            csvs=sorted(d.glob('*.csv'))
            rec={'system':d.name,'readmes':[],'csvs':[]}
            for p in readmes:
                rec['readmes'].append({'path':str(p.relative_to(root)),'sha256':sha256(p),'excerpt':text_excerpt(p,5000)})
            for p in csvs:
                rec['csvs'].append({'path':str(p.relative_to(root)),'sha256':sha256(p),'excerpt':text_excerpt(p,3000)})
            systems.append(rec)
    return {
        'schema':'family-independence-enzyme-db-inventory-v0.1',
        'role':'candidate_generation_only',
        'repository_file_count':len(files),
        'protein_system_count':len(systems),
        'systems':systems,
        'mechanistic_class_assignments':'UNADJUDICATED',
        'admission_decision':'NOT_EVALUATED'
    }

def lipred(root:Path):
    files=manifest(root)
    workbooks=[]
    try:
        import openpyxl
    except Exception:
        openpyxl=None
    for p in sorted(root.rglob('*.xlsx')):
        item={'path':str(p.relative_to(root)),'sha256':sha256(p),'size_bytes':p.stat().st_size}
        if openpyxl:
            try:
                wb=openpyxl.load_workbook(p,read_only=True,data_only=True)
                sheets=[]
                for ws in wb.worksheets:
                    preview=[]
                    for i,row in enumerate(ws.iter_rows(values_only=True)):
                        if i>=25: break
                        preview.append([None if v is None else str(v)[:300] for v in row[:20]])
                    sheets.append({'title':ws.title,'max_row':ws.max_row,'max_column':ws.max_column,'preview':preview})
                item['sheets']=sheets
            except Exception as e:
                item['workbook_error']=str(e)
        workbooks.append(item)
    text_files=[]
    for p in sorted(root.rglob('*')):
        if p.is_file() and p.suffix.lower() in TEXT_EXT and p.stat().st_size < 5_000_000:
            text_files.append({'path':str(p.relative_to(root)),'sha256':sha256(p),'excerpt':text_excerpt(p,8000)})
    return {
        'schema':'family-independence-lipred-inventory-v0.1',
        'role':'candidate_generation_only',
        'repository_file_count':len(files),
        'workbooks':workbooks,
        'text_files':text_files,
        'reaction_family_assignments':'UNADJUDICATED',
        'observed_rate_independence':'NOT_EVALUATED',
        'admission_decision':'NOT_EVALUATED'
    }

def generic(root:Path,label:str):
    return {'schema':'family-independence-generic-inventory-v0.1','label':label,'role':'candidate_generation_only','files':manifest(root),'admission_decision':'NOT_EVALUATED'}

def main():
    ap=argparse.ArgumentParser(); sp=ap.add_subparsers(dest='cmd',required=True)
    for name in ['trypsin','enzyme-db','lipred']:
        s=sp.add_parser(name); s.add_argument('--root',required=True); s.add_argument('--out',required=True)
    s=sp.add_parser('generic'); s.add_argument('--root',required=True); s.add_argument('--label',required=True); s.add_argument('--out',required=True)
    a=ap.parse_args(); root=Path(a.root)
    if not root.is_dir(): raise SystemExit(f'root not found: {root}')
    if a.cmd=='trypsin': obj=trypsin(root)
    elif a.cmd=='enzyme-db': obj=enzyme_db(root)
    elif a.cmd=='lipred': obj=lipred(root)
    else: obj=generic(root,a.label)
    write(a.out,obj)
    print(json.dumps({k:v for k,v in obj.items() if k not in {'files','systems','workbooks','text_files','text_excerpts'}},indent=2,sort_keys=True))

if __name__=='__main__': main()
